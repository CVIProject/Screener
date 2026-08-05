import pandas as pd
import yfinance as yf
import re
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

COLUMN_MAP = {
    "Name, Div % Yield": "Name",
    "Sym": "Symbol",
    "Close": "Price",
    "EPS": "EPS RTG",
    "RS": "REL STR",
    "SECTOR": "Sector"
}


def process_excel(
        input_file,
        biblical_file,
        output_file
):

    print("Reading input Excel...")

    df = pd.read_excel(input_file)
    screen_df = pd.read_excel(biblical_file)

    df.rename(columns=COLUMN_MAP, inplace=True)

    df = df[
        [
            "Name",
            "Symbol",
            "Price",
            "EPS RTG",
            "REL STR",
            "Sector"
        ]
    ].copy()


    df.loc[:, "EPS RTG"] = pd.to_numeric(
        df["EPS RTG"],
        errors="coerce"
    )


    df.loc[:, "REL STR"] = pd.to_numeric(
        df["REL STR"],
        errors="coerce"
    )

    #############################################################
    # Initial Screening
    #############################################################

    standard_filter = (
        (df["EPS RTG"] >= 80) &
        (df["REL STR"] >= 80)
    )

    exception_filter = (
        (df["EPS RTG"] >= 91) &
        (df["REL STR"].between(77, 79))
    )

    exception_filter = (
        (df["REL STR"] >= 91) &
        (df["EPS RTG"].between(77, 79)) 
    )   

    df = df.loc[
        standard_filter | exception_filter
    ].copy()

    print(f"Stocks after EPS/RS filtering : {len(df)}")

    #############################################################
    # Download Historical Prices
    #############################################################

    symbols = (
        df["Symbol"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if len(symbols) == 0:
        df.to_excel(output_file, index=False)
        return

    print("Downloading latest market data from Yahoo Finance...")

    try:

        history = yf.download(
            tickers=symbols,
            period="30mo",
            auto_adjust=True,
            progress=False,
            threads=True,
            group_by="ticker"
        )

    except Exception as ex:
        raise Exception(f"Yahoo Finance Download Failed : {ex}")

    #############################################################
    # Calculate Returns
    #############################################################

    current_price_dict = {}
    six_month_dict = {}
    twentyfour_month_dict = {}


    volatility_dict = {}
    trend_score_dict = {}
    technical_score_dict = {}

    for symbol in symbols:

        try:

            if len(symbols) == 1:
                prices = history["Close"].dropna()

            else:

                if symbol not in history.columns.levels[0]:
                    continue

                prices = history[symbol]["Close"].dropna()

            if prices.empty:
                continue

            latest_price = float(prices.iloc[-1])

            latest_date = prices.index.max()

            six_month_date = latest_date - pd.DateOffset(months=6)
            twentyfour_month_date = latest_date - pd.DateOffset(months=24)

            six_prices = prices.loc[
                prices.index <= six_month_date
            ]

            tf_prices = prices.loc[
                prices.index <= twentyfour_month_date
            ]

            if six_prices.empty or tf_prices.empty:
                continue

            price_6m = float(six_prices.iloc[-1])
            price_24m = float(tf_prices.iloc[-1])

            six_return = (
                (latest_price / price_6m) - 1
            ) * 100

            twentyfour_return = (
                (latest_price / price_24m) - 1
            ) * 100

            current_price_dict[symbol] = round(latest_price, 2)

            six_month_dict[symbol] = round(six_return, 2)

            twentyfour_month_dict[symbol] = round(
                twentyfour_return,
                2
            )


            ########################################################
            # Volatility
            ########################################################

            daily_returns = prices.pct_change().dropna()

            volatility = daily_returns.std() * (252 ** 0.5)

            volatility_dict[symbol] = round(volatility * 100, 2)

            ########################################################
            # Trend Score
            ########################################################

            positive_days = (daily_returns > 0).sum()

            trend_score = (
                positive_days /
                len(daily_returns)
            ) * 100

            trend_score_dict[symbol] = round(trend_score, 2)

            ########################################################
            # Technical Score
            ########################################################

            sma50 = prices.rolling(50).mean().iloc[-1]

            sma200 = prices.rolling(200).mean().iloc[-1]

            score = 0

            if latest_price > sma50:
                score += 1

            if latest_price > sma200:
                score += 1

            if sma50 > sma200:
                score += 1

            technical_score_dict[symbol] = score

        except Exception:
            continue

    #############################################################
    # Populate DataFrame
    #############################################################

    df["Price"] = df["Symbol"].map(current_price_dict)

    df["6-Month Return"] = df["Symbol"].map(
        six_month_dict
    )

    df["24-Month Return"] = df["Symbol"].map(
        twentyfour_month_dict
    )


    df["Volatility"] = df["Symbol"].map(volatility_dict)

    df["Trend Score"] = df["Symbol"].map(trend_score_dict)

    df["Technical Score"] = df["Symbol"].map(technical_score_dict)

    #############################################################
    # Weighted Return
    #############################################################

    df["Weighted Return"] = (
        (df["6-Month Return"] * 0.35)
        +
        (df["24-Month Return"] * 0.65)
    ).round(2)


    df["Final Score"] = (
        0.65 * df["Weighted Return"]
        - 0.20 * df["Volatility"]
        + 0.15 * df["Trend Score"]  
    ).round(2)


    #############################################################
    # Remove Missing Data
    #############################################################

    df.dropna(
    subset=[
        "Price",
        "6-Month Return",
        "24-Month Return",
        "Weighted Return",
        "Volatility",
        "Trend Score",
        "Technical Score",
        "Final Score"
    ],
    inplace=True
)

    #############################################################
    # Additional Screening Rule
    #############################################################

    # Remove stocks where either return is negative
    df = df[
        (df["6-Month Return"] >= 0)
        &
        (df["24-Month Return"] >= 0)
    ].copy()

    # Remove stocks where 6 Month Return > 24 Month Return
    df = df[
        df["6-Month Return"]
        <=
        df["24-Month Return"]
    ].copy()



    df["Overall Rank"] = (
    df["Final Score"]
      .rank(method="dense", ascending=False)
      .astype(int)
    )

    

        #############################################################
    # Rank Within Each Sector
    #############################################################

    # Sort by Sector and Weighted Return
    df.sort_values(
        by=["Sector", "Final Score"],
        ascending=[True, False],
        inplace=True
    )

    # Rank securities inside each sector
    df["Sector Rank"] = (
        df.groupby("Sector")["Final Score"]
        .rank(method="dense", ascending=False)
        .astype(int)
    )

    df.sort_values(
        by=[
            "Sector",
            "Sector Rank"
        ],
        ascending=[
            True,
            True
        ],
        inplace=True
    )





    # =========================================================
    # IBD Sector → S&P 500 Industry Group Mapping
    # Add this after Biblical fail-list filtering
    # =========================================================

    IBD_TO_SP500 = {
        # ENERGY
        "ENERGY": "ENERGY",

        # MATERIALS
        "CHEMICAL": "MATERIALS",
        "METALS": "MATERIALS",
        "MINING": "MATERIALS",

        # REAL ESTATE
        "REAL ESTATE": "REAL ESTATE",

        # INDUSTRIALS
        "BUILDING": "INDUSTRIALS",
        "AEROSPACE": "INDUSTRIALS",
        "AGRICULTURE": "INDUSTRIALS",
        "BUSINESS SVC": "INDUSTRIALS",
        "BUSINESS SER": "INDUSTRIALS",
        "CONSUMER": "INDUSTRIALS",
        "MACHINE": "INDUSTRIALS",
        "OFFICE": "INDUSTRIALS",
        "TRANSPORT": "INDUSTRIALS",

        # CONSUMER DISCRETIONARY
        "APPAREL": "CONSUMER DISCRETIONARY",
        "AUTO": "CONSUMER DISCRETIONARY",
        "LEISURE": "CONSUMER DISCRETIONARY",
        "RETAIL": "CONSUMER DISCRETIONARY",

        # CONSUMER STAPLES
        "FOOD/BEV": "CONSUMER STAPLES",

        # HEALTHCARE
        "MEDICAL": "HEALTHCARE",

        # FINANCIALS
        "BANKS": "FINANCIALS",
        "FINANCE": "FINANCIALS",
        "INSURANCE": "FINANCIALS",
        "S&L": "FINANCIALS",

        # INFORMATION TECHNOLOGY
        "CHIPS": "INFORMATION TECH",
        "COMPUTER": "INFORMATION TECH",
        "SOFTWARE": "INFORMATION TECH",
        "ELECTRONICS": "INFORMATION TECH",

        # COMMUNICATION SERVICES
        "TELECOM": "COMMUNICATION SERVICES",
        "MEDIA": "COMMUNICATION SERVICES",
        "INTERNET": "COMMUNICATION SERVICES",

        # UTILITIES
        "UTILITY": "UTILITIES",

        "MISC": "MISCELLANEOUS"
    }

    # Normalize sector names
    df["Sector"] = (
        df["Sector"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    # Create Industry Group column
    df["Industry Group"] = df["Sector"].map(IBD_TO_SP500)

    # Handle unmapped sectors
    df["Industry Group"] = df["Industry Group"].fillna("UNMAPPED")

    # =========================================================
    # Industry Group Rank
    # =========================================================

    df["Industry Group Rank"] = (
        df.groupby("Industry Group")["Final Score"]
        .rank(method="dense", ascending=False)
        .astype(int)
    )

    # Optional: sort by Industry Group rank
    df.sort_values(
        by=["Industry Group", "Industry Group Rank"],
        ascending=[True, True],
        inplace=True
    )

    df.reset_index(drop=True, inplace=True)

    df.sort_values(
        by=[
            "Industry Group",
            "Industry Group Rank"
        ],
        ascending=[
            True,
            True
        ],
        inplace=True
    )


    #############################################################
    # CLEAN COLUMN NAMES
    #############################################################

    screen_df.columns = (
        screen_df.columns
        .str.strip()
    )

    df.columns = (
        df.columns
        .str.strip()
    )


    #############################################################
    # EXTRACT SYMBOL BEFORE PERIOD
    #############################################################

    def get_base_symbol(symbol):

        if pd.isna(symbol):

            return ""

        symbol = str(
            symbol
        ).strip()


        # Take everything before the first period
        base_symbol = symbol.split(
            ".",
            1
        )[0]


        return base_symbol


    #############################################################
    # PREPARE BIBLICAL FAIL SYMBOLS
    #############################################################

    fail_symbols = (

        screen_df["Symbol"]

        .dropna()

        .apply(get_base_symbol)

        .tolist()

    )


    #############################################################
    # REMOVE EMPTY SYMBOLS
    #############################################################

    fail_symbols = [

        symbol

        for symbol in fail_symbols

        if symbol != ""

    ]


    #############################################################
    # EXACT SYMBOL MATCH FUNCTION
    #############################################################

    def is_biblical_fail_match(

        filtered_symbol,

        fail_symbols

    ):

        if pd.isna(
            filtered_symbol
        ):

            return False


        filtered_symbol = str(

            filtered_symbol

        ).strip()


        if filtered_symbol == "":

            return False


        #########################################################
        # IMPORTANT:
        # DO NOT EXTRACT PREFIX FROM FILTERED SYMBOL
        #
        # The filtered symbol is already:
        #
        # ABT
        # ABBV
        # ACEL
        # ACN
        #########################################################


        for fail_symbol in fail_symbols:


            #####################################################
            # EXACT MATCH ONLY
            #####################################################

            if filtered_symbol == fail_symbol:

                return True


        return False
        #############################################################
    # IDENTIFY BIBLICAL FAIL STOCKS
    #############################################################

    df["Biblical Fail"] = (

        df["Symbol"]

        .apply(

            lambda symbol:

            is_biblical_fail_match(

                symbol,

                fail_symbols

            )

        )

    )


    df["Screen Test"] = "Pass"

    df.loc[
        df["Biblical Fail"] == True,
        "Screen Test"
    ] = "Fail"

        #############################################################
    # SEPARATE PASS AND FAIL STOCKS
    #############################################################

    pass_df = df[
        df["Screen Test"] == "Pass"
    ].copy()

    fail_df = df[
        df["Screen Test"] == "Fail"
    ].copy()


    #############################################################
    # REMOVE HELPER COLUMN
    #############################################################

    pass_df.drop(
        columns=["Biblical Fail"],
        errors="ignore",
        inplace=True
    )

    fail_df.drop(
        columns=["Biblical Fail"],
        errors="ignore",
        inplace=True
    )

    #############################################################
    # CREATE THREE EMPTY ROWS WITHOUT CONCAT WARNING
    #############################################################

    blank_rows = pd.DataFrame(
        pd.NA,
        index=range(3),
        columns=pass_df.columns
)

    #############################################################
    # COMBINE FINAL OUTPUT
    #############################################################
    df = pd.concat(
        [
            pass_df,
            blank_rows,
            fail_df
        ],
        ignore_index=True
    )


    #############################################################
    # FINAL OUTPUT COLUMN ORDER
    #############################################################

    output_columns = [
        "Name",
        "Symbol",
        "Price",
        "Sector",
        "Industry Group",
        "Sector Rank",
        "Industry Group Rank",
        "6-Month Return",
        "24-Month Return",
        "Weighted Return",
        "Volatility",
        "Trend Score",
        "Technical Score",
        "EPS RTG",
        "REL STR",
        "Final Score",
        "Overall Rank",
        "Screen Test"
    ]

    # Keep only columns that exist
    output_columns = [
        column
        for column in output_columns
        if column in df.columns
    ]

    df = df[output_columns]


        #############################################################
    # ADD EMPTY ROW BETWEEN INDUSTRY GROUPS
    #############################################################

    final_rows = []

    previous_group = None

    for _, row in df.iterrows():

        current_group = row["Industry Group"]

        if (
            previous_group is not None
            and current_group != previous_group
        ):
            # Add empty row
            final_rows.append(
                {
                    column: None
                    for column in df.columns
                }
            )

        final_rows.append(row.to_dict())

        previous_group = current_group


    df = pd.DataFrame(final_rows)

   
        #############################################################
    # WRITE DATA TO EXCEL
    #############################################################

    df.to_excel(
        output_file,
        index=False
    )

    #############################################################
    # FORMAT EXCEL FILE
    #############################################################

    wb = load_workbook(output_file)

    ws = wb.active

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header formatting
    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

    # Locate the important columns by header name so inserted columns
    # do not break the formatting logic.
    header_map = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    industry_group_col = header_map.get("Industry Group")
    sector_rank_col = header_map.get("Sector Rank")

    #############################################################
    # LIGHT COLORS FOR INDUSTRY GROUPS
    #############################################################

    industry_colors = [
        "EAF2F8",
        "E8F8F5",
        "FEF9E7",
        "FDEDEC",
        "F4ECF7",
        "E8F6F3",
        "FDF2E9",
        "EBDEF0",
        "EAF2F8",
        "F9E79F",
        "D5F5E3"
    ]

    industry_color_map = {}
    color_index = 0

    for row in range(2, ws.max_row + 1):

        industry_group = None
        if industry_group_col is not None:
            industry_group = ws.cell(
                row=row,
                column=industry_group_col
            ).value

        # Skip empty separator rows
        if industry_group is None:
            continue

        if industry_group not in industry_color_map:
            industry_color_map[industry_group] = (
                industry_colors[
                    color_index % len(industry_colors)
                ]
            )
            color_index += 1

        fill = PatternFill(
            fill_type="solid",
            fgColor=industry_color_map[industry_group]
        )

        for column in range(
            1,
            ws.max_column + 1
        ):
            ws.cell(
                row=row,
                column=column
            ).fill = fill

    #############################################################
    # BOLD TOP 3 STOCKS FROM EACH SECTOR
    #############################################################

    for row in range(
        2,
        ws.max_row + 1
    ):

        sector_rank = None
        if sector_rank_col is not None:
            sector_rank = ws.cell(
                row=row,
                column=sector_rank_col
            ).value

        # Skip empty rows
        if sector_rank is None:
            continue

        try:
            sector_rank = int(sector_rank)

            if sector_rank <= 3:
                for column in range(
                    1,
                    ws.max_column + 1
                ):
                    ws.cell(
                        row=row,
                        column=column
                    ).font = Font(bold=True)

        except (
            ValueError,
            TypeError
        ):
            continue


            #############################################################
    # RED STYLES FOR FAILED STOCKS
    #############################################################

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    red_font = Font(
        color="9C0006"
    )


    #############################################################
    # FIND SCREEN TEST COLUMN
    #############################################################

    screen_test_column = None

    for cell in ws[1]:

        if cell.value == "Screen Test":

            screen_test_column = cell.column

            break


    #############################################################
    # HIGHLIGHT FAILED ROWS
    #############################################################

    if screen_test_column is not None:

        for row in range(
            2,
            ws.max_row + 1
        ):

            screen_test_value = ws.cell(
                row=row,
                column=screen_test_column
            ).value


            if (

                screen_test_value is not None

                and str(
                    screen_test_value
                ).strip().upper() == "FAIL"

            ):

                for column in range(
                    1,
                    ws.max_column + 1
                ):

                    cell = ws.cell(
                        row=row,
                        column=column
                    )

                    cell.fill = red_fill

                    cell.font = red_font

    #############################################################
    # AUTO-SIZE COLUMNS
    #############################################################

    for column_cells in ws.columns:

        max_length = 0

        column_letter = (
            column_cells[0].column_letter
        )

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            30
        )

    wb.save(
    output_file
)

    print("===================================")
    print("Screening Completed Successfully")
    print(f"Stocks Selected : {len(df)}")
    print(f"Output File : {output_file}")
    print("===================================")
