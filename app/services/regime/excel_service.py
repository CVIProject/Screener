from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings


HISTORICAL_COLUMNS = [
    settings.DATE_COLUMN,
    settings.VALUE_COLUMN,
    settings.SLOPE_COLUMN,
    settings.MEDIAN_90_COLUMN,
    settings.MEDIAN_10Y_COLUMN,
    settings.QUAD_COLUMN,
    settings.RULE_3_COLUMN,
    settings.QAD_VALUE_COLUMN,
    settings.RULE_7_COLUMN,
    settings.CONFIRMED_COLUMN,
    settings.TRADE_COLUMN,
]

NEW_DATA_COLUMNS = [
    settings.DATE_COLUMN,
    settings.VALUE_COLUMN,
]

NEW_VALUES_SHEET_NAME = "Daily, Close"


def normalize_column_name(value: object) -> str:
    """
    Normalize an Excel column name by removing surrounding
    whitespace and non-breaking spaces.
    """
    return str(value).replace("\xa0", " ").strip()


def read_new_values_workbook(
    file_content: bytes,
) -> pd.DataFrame:
    """
    Read only the 'Daily, Close' sheet from the new-values workbook.

    Required logical columns:
    - observation_date
    - BAMLH0A0HYM2

    The function also accepts common source column names such as
    Date and Close and renames them to the required application names.
    """
    if not file_content:
        raise ValueError(
            "The new BAML values Excel file is empty."
        )

    try:
        excel_file = pd.ExcelFile(
            BytesIO(file_content),
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            "The new BAML values file is not a readable Excel workbook."
        ) from exc

    if NEW_VALUES_SHEET_NAME not in excel_file.sheet_names:
        available_sheets = ", ".join(excel_file.sheet_names)

        raise ValueError(
            f'The new-values workbook must contain a sheet named '
            f'"{NEW_VALUES_SHEET_NAME}". '
            f"Available sheets: {available_sheets}"
        )

    try:
        dataframe = pd.read_excel(
            BytesIO(file_content),
            sheet_name=NEW_VALUES_SHEET_NAME,
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            f'Unable to read the "{NEW_VALUES_SHEET_NAME}" sheet.'
        ) from exc

    dataframe.columns = [
        normalize_column_name(column)
        for column in dataframe.columns
    ]

    # Support either application headers or common source headers.
    column_aliases = {
        "Date": settings.DATE_COLUMN,
        "DATE": settings.DATE_COLUMN,
        "date": settings.DATE_COLUMN,
        "Observation Date": settings.DATE_COLUMN,
        "observation date": settings.DATE_COLUMN,
        "Close": settings.VALUE_COLUMN,
        "CLOSE": settings.VALUE_COLUMN,
        "close": settings.VALUE_COLUMN,
        "Value": settings.VALUE_COLUMN,
        "VALUE": settings.VALUE_COLUMN,
    }

    dataframe = dataframe.rename(
        columns=column_aliases
    )

    missing_columns = [
        column
        for column in NEW_DATA_COLUMNS
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            f'The "{NEW_VALUES_SHEET_NAME}" sheet is missing '
            f"required columns: {', '.join(missing_columns)}"
        )

    dataframe = dataframe[
        NEW_DATA_COLUMNS
    ].copy()

    dataframe[settings.DATE_COLUMN] = pd.to_datetime(
        dataframe[settings.DATE_COLUMN],
        errors="coerce",
    )

    dataframe[settings.VALUE_COLUMN] = pd.to_numeric(
        dataframe[settings.VALUE_COLUMN],
        errors="coerce",
    )

    # Remove fully blank rows.
    dataframe = dataframe.dropna(
        how="all"
    )

    invalid_mask = (
        dataframe[settings.DATE_COLUMN].isna()
        | dataframe[settings.VALUE_COLUMN].isna()
    )

    if invalid_mask.any():
        invalid_rows = dataframe.index[
            invalid_mask
        ].tolist()

        # Data starts on Excel row 2 because row 1 is the header.
        invalid_excel_rows = [
            str(index + 2)
            for index in invalid_rows
        ]

        raise ValueError(
            f'The "{NEW_VALUES_SHEET_NAME}" sheet contains '
            f"invalid dates or BAMLH0A0HYM2 values at Excel rows: "
            f"{', '.join(invalid_excel_rows)}"
        )

    if (
        dataframe[settings.VALUE_COLUMN] < 0
    ).any():
        negative_rows = dataframe.index[
            dataframe[settings.VALUE_COLUMN] < 0
        ].tolist()

        negative_excel_rows = [
            str(index + 2)
            for index in negative_rows
        ]

        raise ValueError(
            "BAMLH0A0HYM2 values cannot be negative. "
            f"Invalid Excel rows: {', '.join(negative_excel_rows)}"
        )

    dataframe = (
        dataframe
        .sort_values(settings.DATE_COLUMN)
        .drop_duplicates(
            subset=[settings.DATE_COLUMN],
            keep="last",
        )
        .reset_index(drop=True)
    )

    if dataframe.empty:
        raise ValueError(
            f'The "{NEW_VALUES_SHEET_NAME}" sheet contains '
            "no valid observations."
        )

    return dataframe


def find_historical_header_row(
    file_content: bytes,
) -> int:
    """
    Search the first 50 rows of the first worksheet for the
    historical regime header row.
    """
    try:
        preview = pd.read_excel(
            BytesIO(file_content),
            sheet_name=0,
            header=None,
            nrows=50,
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            "Unable to inspect the historical workbook."
        ) from exc

    required_headers = {
        settings.DATE_COLUMN,
        settings.VALUE_COLUMN,
    }

    for row_index in range(len(preview)):
        row_headers = {
            normalize_column_name(value)
            for value in preview.iloc[row_index].tolist()
            if pd.notna(value)
        }

        if required_headers.issubset(
            row_headers
        ):
            return row_index

    raise ValueError(
        "Could not find a historical header row containing "
        f"{settings.DATE_COLUMN} and {settings.VALUE_COLUMN}."
    )


def read_historical_workbook(
    file_content: bytes,
) -> pd.DataFrame:
    """
    Read the historical regime workbook.

    The workbook must contain columns A-K used by the regime module.
    Additional columns are ignored.
    Blank rows are removed.
    """
    if not file_content:
        raise ValueError(
            "The historical regime Excel file is empty."
        )

    try:
        excel_file = pd.ExcelFile(
            BytesIO(file_content),
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            "The historical file is not a readable Excel workbook."
        ) from exc

    if not excel_file.sheet_names:
        raise ValueError(
            "The historical workbook contains no worksheets."
        )

    header_row = find_historical_header_row(
        file_content
    )

    try:
        dataframe = pd.read_excel(
            BytesIO(file_content),
            sheet_name=excel_file.sheet_names[0],
            header=header_row,
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            "Unable to read the historical regime worksheet."
        ) from exc

    dataframe.columns = [
        normalize_column_name(column)
        for column in dataframe.columns
    ]

    missing_columns = [
        column
        for column in HISTORICAL_COLUMNS
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            "The historical workbook is missing required columns: "
            + ", ".join(missing_columns)
        )

    # Select only the required columns.
    # Extra columns such as "7 Day" are ignored.
    dataframe = dataframe[
        HISTORICAL_COLUMNS
    ].copy()

    dataframe[settings.DATE_COLUMN] = pd.to_datetime(
        dataframe[settings.DATE_COLUMN],
        errors="coerce",
    )

    dataframe[settings.VALUE_COLUMN] = pd.to_numeric(
        dataframe[settings.VALUE_COLUMN],
        errors="coerce",
    )

    # Remove blank rows and rows without required input values.
    dataframe = dataframe.dropna(
        subset=[
            settings.DATE_COLUMN,
            settings.VALUE_COLUMN,
        ]
    )

    dataframe = (
        dataframe
        .sort_values(settings.DATE_COLUMN)
        .drop_duplicates(
            subset=[settings.DATE_COLUMN],
            keep="last",
        )
        .reset_index(drop=True)
    )

    if dataframe.empty:
        raise ValueError(
            "The historical workbook contains no valid observations."
        )

    return dataframe


def save_dataframe_to_excel(
    dataframe: pd.DataFrame,
) -> BytesIO:
    """
    Write the complete historical and appended data to a formatted
    Excel workbook.
    """
    result = dataframe.copy()

    result = result[
        HISTORICAL_COLUMNS
    ]

    result[settings.DATE_COLUMN] = (
        pd.to_datetime(
            result[settings.DATE_COLUMN],
            errors="coerce",
        )
        .dt.date
    )

    result[settings.VALUE_COLUMN] = pd.to_numeric(
        result[settings.VALUE_COLUMN],
        errors="coerce",
    )

    result[settings.SLOPE_COLUMN] = (
        pd.to_numeric(
            result[settings.SLOPE_COLUMN],
            errors="coerce",
        )
        .round(3)
    )

    result[settings.MEDIAN_90_COLUMN] = (
        pd.to_numeric(
            result[settings.MEDIAN_90_COLUMN],
            errors="coerce",
        )
        .round(2)
    )

    result[settings.MEDIAN_10Y_COLUMN] = (
        pd.to_numeric(
            result[settings.MEDIAN_10Y_COLUMN],
            errors="coerce",
        )
        .round(2)
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
        date_format="yyyy-mm-dd",
        datetime_format="yyyy-mm-dd",
    ) as writer:
        result.to_excel(
            writer,
            index=False,
            sheet_name="Regime Data",
        )

        worksheet = writer.sheets[
            "Regime Data"
        ]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):
            worksheet.cell(
                row=row_number,
                column=1,
            ).number_format = "yyyy-mm-dd"

            worksheet.cell(
                row=row_number,
                column=2,
            ).number_format = "0.00"

            worksheet.cell(
                row=row_number,
                column=3,
            ).number_format = "0.000"

            worksheet.cell(
                row=row_number,
                column=4,
            ).number_format = "0.00"

            worksheet.cell(
                row=row_number,
                column=5,
            ).number_format = "0.00"

        # Hide intermediate rule columns if required.
        worksheet.column_dimensions[
            "G"
        ].hidden = True

        worksheet.column_dimensions[
            "H"
        ].hidden = True

        worksheet.column_dimensions[
            "I"
        ].hidden = True

        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            maximum_length = max(
                (
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )

            column_letter = get_column_letter(
                column_number
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                50,
            )

    output.seek(0)

    return output