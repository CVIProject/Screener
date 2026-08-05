from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.consensus.models import ConsensusConfig, ConsensusResult


NAVY = "1F4E78"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
DARK_GREEN = "548235"
LIGHT_GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
GRAY = "E7E6E6"
PORTFOLIO_BLUE = "BDD7EE"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color=WHITE, bold=True)
SECTOR_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
INDUSTRY_FILL = PatternFill("solid", fgColor=PALE_BLUE)
DARK_GREEN_FILL = PatternFill("solid", fgColor=DARK_GREEN)
LIGHT_GREEN_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
RED_FILL = PatternFill("solid", fgColor=RED)
GRAY_FILL = PatternFill("solid", fgColor=GRAY)
PORTFOLIO_FILL = PatternFill("solid", fgColor=PORTFOLIO_BLUE)

THIN_SIDE = Side(style="thin", color="D9E2F3")
MEDIUM_SIDE = Side(style="medium", color="7F8C8D")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)



def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = THIN_BORDER


def _date_header(week) -> str:
    """
    Display the date detected from the filename as MM-DD-YY.

    Example filename:
        screening_10-22-2025.xlsx

    Output header:
        10-22-25
    """
    if week.week_date is not None:
        return week.week_date.strftime("%m-%d-%y")

    # No date was detected in the filename.
    return week.week_label


def _weekly_lookup(
    result: ConsensusResult,
) -> dict[tuple[str, str, int, int], dict]:
    lookup: dict[tuple[str, str, int, int], dict] = {}

    for _, row in result.weekly_top_five.iterrows():
        if pd.isna(row.get("weekly_rank")):
            continue

        lookup[
            (
                str(row["sector"]),
                str(row["industry"]),
                int(row["week_order"]),
                int(row["weekly_rank"]),
            )
        ] = row.to_dict()

    return lookup


def _consensus_lookup(
    result: ConsensusResult,
) -> dict[tuple[str, str], list[dict]]:
    lookup: dict[tuple[str, str], list[dict]] = {}

    ordered = result.consensus_top_five.sort_values(
        ["sector", "industry", "industry_consensus_rank"]
    )

    for key, group in ordered.groupby(
        ["sector", "industry"],
        sort=False,
    ):
        lookup[(str(key[0]), str(key[1]))] = group.to_dict("records")

    return lookup


def _portfolio_lookup(
    result: ConsensusResult,
) -> dict[tuple[str, str], list[dict]]:
    """
    Group every portfolio holding by its detected sector and industry.

    Holdings are included even when they are not part of the final top five.
    """
    lookup: dict[tuple[str, str], list[dict]] = {}

    if (
        result.portfolio_comparison is None
        or result.portfolio_comparison.empty
    ):
        return lookup

    valid = result.portfolio_comparison[
        result.portfolio_comparison["sector"].notna()
        & result.portfolio_comparison["industry"].notna()
    ].copy()

    valid = valid.sort_values(
        [
            "sector",
            "industry",
            "portfolio_industry_rank",
            "portfolio_name",
            "ticker",
        ],
        na_position="last",
    )

    for key, group in valid.groupby(
        ["sector", "industry"],
        sort=False,
    ):
        lookup[(str(key[0]), str(key[1]))] = group.to_dict("records")

    return lookup


def _action_fill(action: str) -> PatternFill:
    action = str(action).upper()

    if action == "HOLD / ADD":
        return DARK_GREEN_FILL
    if action == "HOLD":
        return LIGHT_GREEN_FILL
    if action.startswith("REPLACE WITH"):
        return RED_FILL
    if "NO SCREENING DATA" in action:
        return GRAY_FILL
    return YELLOW_FILL


def build_workbook(
    result: ConsensusResult,
    config: ConsensusConfig,
) -> bytes:
    """
    Generate exactly one simple sheet.

    Every industry block contains:
    - all portfolio holdings assigned to that industry,
    - each holding's rank versus the final industry top five,
    - weekly top five company/ticker pairs,
    - final top five company/ticker pairs,
    - HOLD / ADD, HOLD, or REPLACE WITH <ticker>.
    """
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Top 5 Comparison"

    weekly_lookup = _weekly_lookup(result)
    consensus_lookup = _consensus_lookup(result)
    portfolio_lookup = _portfolio_lookup(result)

    # Stocks appearing in more than 4 uploaded filtered files.
    common_stock_tickers = set(
        result.stock_summary.loc[
            result.stock_summary["frequency"] > 4,
            "ticker",
        ]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    # Build complete hierarchy from final top five and portfolio industries.
    hierarchy = set(consensus_lookup.keys()) | set(portfolio_lookup.keys())
    hierarchy = sorted(hierarchy, key=lambda item: (item[0], item[1]))

    column = 1
    sector_col = column
    column += 1
    industry_col = column
    column += 1

    portfolio_model_col = column
    portfolio_company_col = column + 1
    portfolio_ticker_col = column + 2
    portfolio_rank_col = column + 3
    column += 4

    week_columns: list[tuple[int, int, int]] = []
    for week in result.weekly_screens:
        week_columns.append((week.week_order, column, column + 1))
        column += 2

    final_company_col = column
    final_ticker_col = column + 1
    column += 2

    action_col = column
    last_col = action_col

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_col,
    )
    ws.cell(1, 1, "Portfolio vs Industry Top 5 Comparison")
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).font = Font(
        color=WHITE,
        bold=True,
        size=16,
    )
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells(
        start_row=2,
        start_column=sector_col,
        end_row=3,
        end_column=sector_col,
    )
    ws.cell(2, sector_col, "Sector")

    ws.merge_cells(
        start_row=2,
        start_column=industry_col,
        end_row=3,
        end_column=industry_col,
    )
    ws.cell(2, industry_col, "Industry")

    ws.merge_cells(
        start_row=2,
        start_column=portfolio_model_col,
        end_row=2,
        end_column=portfolio_rank_col,
    )
    ws.cell(2, portfolio_model_col, "Portfolio Stock Comparison")
    ws.cell(3, portfolio_model_col, "Model")
    ws.cell(3, portfolio_company_col, "Portfolio Stock")
    ws.cell(3, portfolio_ticker_col, "Ticker")
    ws.cell(3, portfolio_rank_col, "Industry Rank")

    for week, (_, company_col, ticker_col) in zip(
        result.weekly_screens,
        week_columns,
    ):
        ws.merge_cells(
            start_row=2,
            start_column=company_col,
            end_row=2,
            end_column=ticker_col,
        )
        # Direct MM-DD-YY date; no "Week Ending" text.
        ws.cell(2, company_col, _date_header(week))
        ws.cell(3, company_col, "Company")
        ws.cell(3, ticker_col, "Ticker")

    ws.merge_cells(
        start_row=2,
        start_column=final_company_col,
        end_row=2,
        end_column=final_ticker_col,
    )
    ws.cell(2, final_company_col, "Final Industry Top 5")
    ws.cell(3, final_company_col, "Company")
    ws.cell(3, final_ticker_col, "Ticker")

    ws.merge_cells(
        start_row=2,
        start_column=action_col,
        end_row=3,
        end_column=action_col,
    )
    ws.cell(2, action_col, "Comparison Action")

    for row in (2, 3):
        for col in range(1, last_col + 1):
            _style_header(ws.cell(row, col))

    current_row = 4

    for sector, industry in hierarchy:
        holdings = portfolio_lookup.get((sector, industry), [])
        final_top = consensus_lookup.get((sector, industry), [])

        # Always preserve five rows for final top five, but expand when an
        # industry contains more than five portfolio holdings.
        block_size = max(
            config.top_n_per_industry,
            len(holdings),
        )
        start_row = current_row
        end_row = start_row + block_size - 1

        ws.merge_cells(
            start_row=start_row,
            start_column=sector_col,
            end_row=end_row,
            end_column=sector_col,
        )
        ws.merge_cells(
            start_row=start_row,
            start_column=industry_col,
            end_row=end_row,
            end_column=industry_col,
        )

        ws.cell(start_row, sector_col, sector)
        ws.cell(start_row, industry_col, industry)
        ws.cell(start_row, sector_col).fill = SECTOR_FILL
        ws.cell(start_row, industry_col).fill = INDUSTRY_FILL
        ws.cell(start_row, sector_col).font = Font(bold=True)
        ws.cell(start_row, industry_col).font = Font(bold=True)

        for col in (sector_col, industry_col):
            ws.cell(start_row, col).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for offset in range(block_size):
            row = start_row + offset

            for col in range(1, last_col + 1):
                ws.cell(row, col).border = THIN_BORDER
                ws.cell(row, col).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            # All portfolio holdings for this industry, even when absent from
            # the final top five.
            if offset < len(holdings):
                holding = holdings[offset]

                portfolio_name = holding.get("portfolio_name")
                company = holding.get("company_portfolio")
                if pd.isna(company) or not str(company).strip():
                    company = holding.get("company")
                if pd.isna(company) or not str(company).strip():
                    company = holding.get("ticker")

                ws.cell(row, portfolio_model_col, portfolio_name)
                ws.cell(row, portfolio_company_col, company)
                ws.cell(row, portfolio_ticker_col, holding.get("ticker"))
                ws.cell(
                    row,
                    portfolio_rank_col,
                    holding.get("portfolio_industry_rank"),
                )

                action = str(holding.get("portfolio_action", "REVIEW"))
                ws.cell(row, action_col, action)

                for col in (
                    portfolio_model_col,
                    portfolio_company_col,
                    portfolio_ticker_col,
                    portfolio_rank_col,
                ):
                    ws.cell(row, col).fill = PORTFOLIO_FILL

                ws.cell(row, action_col).fill = _action_fill(action)

                if action == "HOLD / ADD":
                    ws.cell(row, action_col).font = Font(
                        color=WHITE,
                        bold=True,
                    )
                elif action.startswith("REPLACE WITH"):
                    ws.cell(row, action_col).font = Font(
                        color="9C0006",
                        bold=True,
                    )

            # Weekly top five occupies the first five rows.
            top_position = offset + 1
            if top_position <= config.top_n_per_industry:
                for week_order, company_col, ticker_col in week_columns:
                    item = weekly_lookup.get(
                        (
                            sector,
                            industry,
                            week_order,
                            top_position,
                        )
                    )
                    if item:
                        company = item.get("company")
                        ticker = str(item.get("ticker", "")).strip().upper()

                        company_cell = ws.cell(row, company_col, company)
                        ticker_cell = ws.cell(row, ticker_col, ticker)

                        # Highlight common stocks in every week where they appear.
                        if ticker in common_stock_tickers:
                            company_cell.fill = YELLOW_FILL
                            ticker_cell.fill = YELLOW_FILL

                            company_cell.font = Font(
                                color="9C6500",
                                bold=True,
                            )
                            ticker_cell.font = Font(
                                color="9C6500",
                                bold=True,
                            )

                if offset < len(final_top):
                    stock = final_top[offset]
                    ws.cell(
                        row,
                        final_company_col,
                        stock.get("company"),
                    )
                    ws.cell(
                        row,
                        final_ticker_col,
                        stock.get("ticker"),
                    )

                    if stock.get("recommendation") == "High Conviction":
                        ws.cell(
                            row,
                            final_company_col,
                        ).fill = DARK_GREEN_FILL
                        ws.cell(
                            row,
                            final_ticker_col,
                        ).fill = DARK_GREEN_FILL
                        ws.cell(
                            row,
                            final_company_col,
                        ).font = Font(color=WHITE, bold=True)
                        ws.cell(
                            row,
                            final_ticker_col,
                        ).font = Font(color=WHITE, bold=True)
                    else:
                        ws.cell(
                            row,
                            final_company_col,
                        ).fill = LIGHT_GREEN_FILL
                        ws.cell(
                            row,
                            final_ticker_col,
                        ).fill = LIGHT_GREEN_FILL

        for col in range(1, last_col + 1):
            cell = ws.cell(end_row, col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=MEDIUM_SIDE,
            )

        current_row = end_row + 1

    ws.column_dimensions[get_column_letter(sector_col)].width = 24
    ws.column_dimensions[get_column_letter(industry_col)].width = 22
    ws.column_dimensions[get_column_letter(portfolio_model_col)].width = 18
    ws.column_dimensions[get_column_letter(portfolio_company_col)].width = 20
    ws.column_dimensions[get_column_letter(portfolio_ticker_col)].width = 10
    ws.column_dimensions[get_column_letter(portfolio_rank_col)].width = 14

    for _, company_col, ticker_col in week_columns:
        ws.column_dimensions[get_column_letter(company_col)].width = 20
        ws.column_dimensions[get_column_letter(ticker_col)].width = 10

    ws.column_dimensions[get_column_letter(final_company_col)].width = 22
    ws.column_dimensions[get_column_letter(final_ticker_col)].width = 10
    ws.column_dimensions[get_column_letter(action_col)].width = 28

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 34

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = (
        f"A3:{get_column_letter(last_col)}{current_row - 1}"
    )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def output_filename() -> str:
    return (
        f"portfolio_vs_industry_top5_"
        f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
